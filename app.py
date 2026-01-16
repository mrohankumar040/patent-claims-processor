import streamlit as st
import pandas as pd
import openpyxl as ox
from datetime import datetime
import re
import io
from pathlib import Path


def remove_non_ascii(s):
    return "".join(filter(lambda x: ord(x) < 128, s)) if isinstance(s, str) else s


def clean_claim_text(claim):
    if not claim or not isinstance(claim, str):
        return "", []

    text = claim.strip()
    text = remove_non_ascii(text)
    comments = []

    if "(canceled)" in text.lower():
        comments.append("cancel")
    if any(x in text.lower() for x in ["non-transitory computer-readable medium", "computer-readable medium", "computer program", "computer"]):
        comments.append("CRM")
    if "delete" in text.lower():
        comments.append("delete")
    if "paragraph" in text.lower():
        comments.append("paragraph")
    if "article" in text.lower():
        comments.append("article")
    if "clause" in text.lower():
        comments.append("clause")

    if re.search(r'\bclaims?\b\s*\d+', text, re.IGNORECASE):
        comments.append("probable dependent")

    if "probable dependent" not in comments and len(text) < 200:
        comments.append("probable dependent")

    return text, comments


def process_details_sheet(details_file):
    """Process the details sheet and return the cleaned workbook as bytes"""
    wb = ox.load_workbook(details_file)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    claim_col_idx = headers.index("Independent Claims")
    key_col_idx = 0

    updated_rows = [headers]
    rejected_rows = [["Key", "Claim", "Comments"]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_list = list(row)
        key = row[key_col_idx]
        claim_cell = row[claim_col_idx]
        cleaned_claims = []
        if claim_cell:
            parts = str(claim_cell).split('|')
            for part in parts:
                cleaned_text, comments = clean_claim_text(part)
                if any(c in comments for c in ["cancel", "CRM", "delete", "paragraph", "article", "clause"]):
                    rejected_rows.append([key, cleaned_text, ', '.join(comments)])
                else:
                    cleaned_claims.append(cleaned_text)
        row_list[claim_col_idx] = ' | '.join(cleaned_claims)
        updated_rows.append(row_list)

    out_wb = ox.Workbook()
    main_ws = out_wb.active
    main_ws.title = "Updated Details"
    for row in updated_rows:
        main_ws.append(row)

    rejected_ws = out_wb.create_sheet(title="Commented Claims")
    for row in rejected_rows:
        rejected_ws.append(row)

    # Save to BytesIO
    output = io.BytesIO()
    out_wb.save(output)
    output.seek(0)
    return output


def process_info_sheet(metadata_file, family_published_file, cleaned_details_file):
    """Process info sheet and return as bytes"""
    df_meta = pd.read_excel(metadata_file)
    df_family_published = pd.read_excel(family_published_file)
    df_cleaned = pd.read_excel(cleaned_details_file)

    # Get seed patents only
    df_seed = df_meta[df_meta['Seed?'] == True].copy()

    # Create claim map
    claim_map = dict(zip(df_cleaned['Publication Number'], df_cleaned['Independent Claims']))

    # Base info sheet
    df_info = pd.DataFrame({
        'Publication Number': df_seed['Seed patent'],
        'Application Number': df_seed['Formatted Application Number'],
        'INPADOC Family Members': df_seed['extended family members'].astype(str).fillna(''),
        'Claim Number': df_seed['Claim number']
    })

    # Clean 'Claim Number' column
    df_info['Claim Number'] = df_info['Claim Number'].astype(str).str.extract(r'(\d+)', expand=False).astype('Int64')

    # Map Family ID
    df_family_lookup = df_family_published[['Publication Number', 'Simple Family Id']]
    df_info = df_info.merge(df_family_lookup, on='Publication Number', how='left')
    df_info.rename(columns={'Simple Family Id': 'INPADOC Family ID'}, inplace=True)

    # Get earliest publication per Application Number
    df_family_published['Publication Date'] = pd.to_datetime(df_family_published['Publication Date'], errors='coerce')
    df_sorted = df_family_published.sort_values(['Application Number', 'Publication Date'], ascending=[True, True])
    df_earliest_pub = df_sorted.drop_duplicates(subset='Application Number', keep='first')
    earliest_pub_map = dict(zip(df_earliest_pub['Application Number'], df_earliest_pub['Publication Number']))

    def concat_published(row):
        existing = row['INPADOC Family Members']
        earliest_pub = earliest_pub_map.get(row['Application Number'])

        if not earliest_pub:
            return existing

        existing_set = set(existing.split('|')) if existing else set()

        if earliest_pub not in existing_set:
            existing_set.add(earliest_pub)

        return "|".join(sorted(existing_set))

    df_info['INPADOC Family Members'] = df_info.apply(concat_published, axis=1)

    # Final columns
    df_info = df_info[
        ['Publication Number', 'Application Number', 'INPADOC Family ID',
         'INPADOC Family Members', 'Claim Number']
    ]

    # Save to BytesIO
    output = io.BytesIO()
    seed_count = len(df_seed)
    df_info.to_excel(output, index=False)
    output.seek(0)
    return output, seed_count


# Streamlit UI
st.set_page_config(page_title="Patent Claims Processor", page_icon="📄", layout="wide")

st.title("📄 Patent Claims Processor")
st.markdown("### Process your patent claim data files easily")
st.markdown("---")

# Instructions
with st.expander("📖 Instructions - Click to expand"):
    st.markdown("""
    **Required Files:**
    1. **Details Sheet** - Must contain "Independent Claims" column and family publication data
    2. **Metadata Sheet** - Must contain seed patent information (with "metadata" in filename)
    
    **What this tool does:**
    - Cleans and filters patent claims from the Details sheet
    - Removes canceled, CRM, and dependent claims
    - Processes family publication data for INPADOC lookups
    - Generates two output files:
        - Details sheet CC.xlsx (cleaned claims with "Commented Claims" sheet)
        - Info_Sheet_CC_[count].xlsx (processed info sheet with family data)
    """)

st.markdown("---")

# File uploaders
col1, col2 = st.columns(2)

with col1:
    st.subheader("1️⃣ Details Sheet")
    details_file = st.file_uploader("Upload Details Sheet", type=['xlsx'], key="details")

with col2:
    st.subheader("2️⃣ Metadata Sheet")
    metadata_file = st.file_uploader("Upload Metadata Sheet", type=['xlsx'], key="metadata")

st.markdown("---")

# Process button
if st.button("🚀 Process Files", type="primary", use_container_width=True):
    if not all([details_file, metadata_file]):
        st.error("⚠️ Please upload both required files!")
    else:
        try:
            with st.spinner("Processing your files... Please wait."):
                # Process details sheet
                cleaned_details_output = process_details_sheet(details_file)
                
                # Process info sheet - using details_file for both purposes
                info_output, seed_count = process_info_sheet(
                    metadata_file, 
                    details_file,  # Using the same details file for family data
                    cleaned_details_output
                )
                
                # Reset the BytesIO for download
                cleaned_details_output.seek(0)
                
            st.success("✅ Processing completed successfully!")
            
            # Download buttons
            st.markdown("### 📥 Download Your Processed Files")
            col1, col2 = st.columns(2)
            
            with col1:
                st.download_button(
                    label="📄 Download Details Sheet CC.xlsx",
                    data=cleaned_details_output,
                    file_name="Details_sheet_CC.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col2:
                st.download_button(
                    label="📊 Download Info Sheet CC.xlsx",
                    data=info_output,
                    file_name=f"Info_Sheet_CC_{seed_count}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
        except Exception as e:
            st.error(f"❌ Error processing files: {str(e)}")
            st.info("Please check that your files have the correct format and required columns.")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: gray; padding: 20px;'>
    <p>Patent Claims Processor v2.1 | Built with Streamlit</p>
</div>
""", unsafe_allow_html=True)
